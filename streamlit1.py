import streamlit as st
import pandas as pd
import numpy as np
import datetime as dt
today=dt.datetime.now()
from pkg_resources import resource_filename
#import plotly.express as px
#import seaborn as sns
#import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

st.title('Pronósticos')

# data_file = st.file_uploader("Upload XLSX", type=["XLSX"]) # Diseño

# if data_file is not None:
    
#     df_input = pd.read_excel(r"Planes.xlsx", sheet_name="Hoja1") # Se importan datos
   

file = st.file_uploader("Cargar archivo de Excel", type=["xlsx"])

if file:
    # Leer el archivo Excel
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # Leer datos desde el archivo Excel
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Mostrar los datos en Streamlit
    st.table(data)

#traemos los datos
#datos = pd.read_excel(r"Planes.xlsx")

#datos = pd.read_excel('C:/Users/ACjdpino/Desktop/Pronósticos panama/Pronosticos colombia/Visualización/Planes.xlsx')
#materiales=datos['Material'].drop_duplicates().tolist() #creamos lista de materiales
#sel_mes=st.select_slider('Mes',options=datos['Mes'].drop_duplicates() )
#sel_sem=st.select_slider('Semana',options=datos['Año natural/Semana'].drop_duplicates())
#sel_mat=st.selectbox('Material', materiales)#datos['Material'].drop_duplicates)


#funcion para mostrar el dataframe que se cargo 
#def mostrar_tabla(df):
#st.write(df)

#fig = px.line(datos, x=datos['Año natural/Semana'],y=datos['Ajuste Plan final Línea'], color=datos['tipo_dato'])

#st.plotly_chart(fig)
#sns.set(style="darkgrid")

#sns.lineplot(data=datos,x='Año natural/Semana',y='Ajuste Plan final Línea',hue=datos['tipo_dato'],sort=True,style=datos['tipo_dato'],errorbar=None,palette=['purple','blue'], markers=True,)

#sns.mpl.pyplot.show()
#st.pyplot()

#edit_df=st.data_editor(datos,disabled=['Año natural/Semana','Mes','Año natural/Mes','Material','DescMat','Centro','DescCe','tipo_dato']
 #                      ,hide_index=True,width=1080, height=400)

#if st.button("Guardar cambios"):
 #   edited_df=edit_df.copy()
    #new_df.to_excel('uno.xlsx', index=False)
  #  st.write('Cambios guardados🎈')
